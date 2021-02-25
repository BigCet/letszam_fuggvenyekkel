from fuggvenyeim.rutinok import honap_atalakito, napok_validalas, szakok_validalasa
from openpyxl import Workbook, load_workbook

honap = honap_atalakito()

nap = napok_validalas()

szak = szakok_validalasa()

elrend_ok = input("Elrendelés oka, ha nem írsz be semmit, akkor létszámpótlás:" )

if elrend_ok == "":
    elrend_ok = "Létszámpótlás"

en_muszakom = 1
filename_nap = nap

ADAT_FILE = f"MŰSZAKNAPLÓ/Műszaknapló_{szak}.xlsx"
TULORA_FILE = "MŰSZAKNAPLÓ/tuloracsoportos.xlsx"

munkafuzet = Workbook()
muszaknaplo = load_workbook(filename=ADAT_FILE)
munkafuzet_lap = munkafuzet.active
tulora = load_workbook(filename=TULORA_FILE)
lap = tulora.active

lap["C3"] = f"2021.{honap}"
lap["I5"] = f"Szervezet: Termelési Osztály {szak} szak."
lap["A6"] = f"Elrendelés oka: {elrend_ok}. "

oszlop = nap + 4

a = 3
b = 10
c = -1
vege = 59

sorszam = 0

for row in range(a, vege):
    a = a + 1
    aktualis_lap = muszaknaplo[f"2021 {honap}"]

    if type(aktualis_lap.cell(row=a, column=oszlop).value) != int:
        continue

    b = b + 1
    c = c + 1

    tsz = munkafuzet_lap[f"B{b}"] = aktualis_lap[f"B{a}"].value
    name = munkafuzet_lap[f"C{b}"] = aktualis_lap[f"C{a}"].value
    en_muszakom = aktualis_lap.cell(row=a, column=oszlop).value

    tol = 6
    ig = 14
    ora = 8
    elrend_cod = 2
    elrend_cod2 = 1

    if en_muszakom == 1:
        tol = 6
        ig = 14
    elif en_muszakom == 2:
        tol = 14
        ig = 22
    elif en_muszakom == 3:
        tol = 22
        ig = 6

    uj_adat = [tsz, name, nap, tol, ig, 8, 2, "", "", "", "", 1]

    if c == 25:

        tulora.save(f"tulora_{szak}_{honap}_{filename_nap}_{sorszam}.xlsx")

        sorszam = sorszam + 1
        c = 0
        b = 11

    for i in lap[f"B{b}": f"M{b}"]:
        for index, cell in enumerate(i):
            cell.value = uj_adat[index]
        print(c, b, )

for row in range(b, 35):
    b = b + 1

    tsz = munkafuzet_lap[f"B{b}"] = ""
    name = munkafuzet_lap[f"C{b}"] = ""
    nap = munkafuzet_lap[f"D{b}"] = ""
    tol = munkafuzet_lap[f"E{b}"] = ""
    ig = munkafuzet_lap[f"F{b}"] = ""
    ora = munkafuzet_lap[f"G{b}"] = ""
    elrend_cod = munkafuzet_lap[f"H{b}"] = ""
    elrend_cod2 = munkafuzet_lap[f"M{b}"] = ""

    uj_adat = ["", name, nap, tol, ig, ora, elrend_cod, "", "", "", "", elrend_cod2]

    for i in lap[f"B{b}": f"M{b}"]:
        for index, cell in enumerate(i):
            cell.value = uj_adat[index]
            print(c, b)

tulora.save(f"tulora_{szak}_{honap}_{filename_nap}_{sorszam}.xlsx")
